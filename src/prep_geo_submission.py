#!/usr/bin/env python3
"""
Prepare GEO submission metadata for Parse Biosciences scRNA-seq experiment.

Reads project metadata, FASTQ inventory, and processed output paths, then
writes a filled GEO metadata Excel file based on the GEO template structure.

Usage:
    python src/prep_geo_submission.py [--output OUTPUT]
"""
import argparse
import csv
import hashlib
import os
import re
import shutil
import sys
from pathlib import Path

import openpyxl
import yaml

PROJECT_DIR = Path(__file__).resolve().parent.parent
CONFIG_PATH = PROJECT_DIR / "config.yaml"
METADATA_CSV = PROJECT_DIR / "metadata" / "metadata.csv"
FASTQ_DIR = PROJECT_DIR / "data" / "FASTQ"
CHECKSUMS_FILE = PROJECT_DIR / "data" / "checksums.md5"
SAMPLE_LIST_FILE = PROJECT_DIR / "data" / "sample_list.txt"
PROCESSED_DIR = PROJECT_DIR / "output" / "parse_comb"
TEMPLATE_PATH = PROJECT_DIR / "docs" / "parse_template.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_DIR / "output"

# ---- Experiment-level constants ----
ORGANISM = "Homo sapiens"
LIBRARY_STRATEGY = "scRNA-seq"
MOLECULE = "polyA RNA"
SINGLE_OR_PAIRED = "paired-end"
INSTRUMENT_MODEL = "Illumina NovaSeq X Plus"
GENOME_BUILD = "GRCh38 (hg38)"

EXTRACT_PROTOCOL = (
    "Human embryonic stem cells (hESCs) were differentiated toward an osteogenic "
    "lineage for 7 or 20 days. At the designated time point, cells were exposed to "
    "the indicated toxicant (bisphenol analogue or Zyn nicotine pouch extract) at the "
    "indicated concentration (IC10, IC25, or IC50) or left untreated. Cells were then "
    "fixed and processed following the Parse Biosciences Whole Transcriptome Kit (WTK) "
    "fixation protocol."
)

LIBRARY_CONSTRUCTION_PROTOCOL = (
    "Single-cell RNA-seq libraries were prepared using the Parse Biosciences Whole "
    "Transcriptome Kit (WT Mega, v3 chemistry) following the manufacturer's protocol. "
    "Fixed cells from all 30 samples were combinatorially barcoded and pooled across "
    "8 sublibraries for sequencing."
)

DATA_PROCESSING_STEPS = [
    (
        "Single-cell libraries were generated using the Parse Biosciences combinatorial "
        "barcoding platform (8 sublibraries sequenced across 3 lanes). All samples were "
        "multiplexed across the shared raw FASTQ files; individual sample demultiplexing "
        "and expression quantification were performed by the split-pipe pipeline based on "
        "cell barcode assignments. The same pooled FASTQ files are therefore listed as the "
        "raw files for every sample."
    ),
    (
        "Raw FASTQ files from each sublibrary were processed independently using "
        "split-pipe v1.6.0 (Parse Biosciences) in '--mode all' with the GRCh38 (hg38) "
        "reference transcriptome."
    ),
    (
        "Per-sublibrary outputs were combined using split-pipe v1.6.0 in '--mode comb' "
        "to generate per-sample demultiplexed DGE matrices."
    ),
    (
        "Processed data files per sample: count_matrix.mtx (sparse UMI count matrix in "
        "Matrix Market format), all_genes.csv (gene ID and gene name annotations), "
        "cell_metadata.csv (per-cell barcode and QC metadata)."
    ),
]

PROCESSED_FILES_FORMAT = (
    "count_matrix.mtx: sparse UMI count matrix (Matrix Market format, cells × genes); "
    "all_genes.csv: gene annotations (Ensembl ID, gene name); "
    "cell_metadata.csv: per-cell metadata including barcode, sample assignment, and QC metrics."
)

STUDY_SUMMARY_PLACEHOLDER = (
    "[TODO: Insert abstract here. Describe the study goal, experimental system "
    "(hESC osteogenic differentiation), toxicants tested (BPA, BPF, BPS; Zyn pouches), "
    "and key findings.]"
)

EXPERIMENTAL_DESIGN_PLACEHOLDER = (
    "Human embryonic stem cells were differentiated toward osteoblasts over 20 days. "
    "Cells were treated with bisphenol analogues (BPA, BPF, BPS) at IC10/IC25/IC50 "
    "concentrations at day 7 or day 20, or with Zyn nicotine pouch extracts "
    "(Chill, Wintergreen, Smooth, Original; 3 mg or 6 mg) at day 7. "
    "Untreated controls were collected at day 0, 7, and 20. Single-cell RNA-seq "
    "was performed using Parse Biosciences combinatorial barcoding across 8 "
    "sublibraries (30 samples, ~100,000 cells total)."
)


def load_metadata(path: Path) -> list[dict]:
    with open(path, newline="") as fh:
        return list(csv.DictReader(fh))


def load_sample_list(path: Path) -> list[tuple[str, str]]:
    """Return (sample_name, full_line) for each entry of the Parse sample list.

    Each line is '<sample_name> <well(s)>'; the full line is preserved verbatim
    so the per-experiment lists are exact subsets of the original.
    """
    entries: list[tuple[str, str]] = []
    with open(path) as fh:
        for line in fh:
            text = line.rstrip("\n")
            if text.strip():
                entries.append((text.split()[0], text))
    return entries


def load_submission_titles(path: Path) -> dict[str, str]:
    """Return {experiment_id: title} from the `geo.submissions` config block."""
    with open(path) as fh:
        config = yaml.safe_load(fh) or {}
    submissions = (config.get("geo") or {}).get("submissions") or {}
    return {str(exp): entry["title"] for exp, entry in submissions.items()}


def load_checksums(path: Path) -> dict[str, str]:
    """Return {filename_basename: md5} from a standard md5sum file."""
    checksums: dict[str, str] = {}
    if not path.exists():
        return checksums
    with open(path) as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            md5, relpath = line.split(None, 1)
            checksums[Path(relpath).name] = md5
    return checksums


def collect_fastq_runs(fastq_dir: Path) -> list[dict]:
    """
    Return list of dicts with keys: sublibrary, lane, R1, R2, I1, I2.
    Derived from the FASTQ directory structure.
    """
    runs: list[dict] = []
    for sub_dir in sorted(fastq_dir.iterdir()):
        if not sub_dir.is_dir():
            continue
        sublibrary = sub_dir.name
        lane_files: dict[str, dict[str, str]] = {}
        for fq in sorted(sub_dir.glob("*.fastq.gz")):
            m = re.search(r"_(L\d+)_(I1|I2|R1|R2)_\d+\.fastq\.gz$", fq.name)
            if not m:
                continue
            lane, read = m.group(1), m.group(2)
            lane_files.setdefault(lane, {})[read] = fq.name
        for lane, files in sorted(lane_files.items()):
            # Only emit rows that have at least R1 and R2
            if "R1" not in files or "R2" not in files:
                continue
            runs.append(
                {
                    "sublibrary": sublibrary,
                    "lane": lane,
                    "R1": files.get("R1", ""),
                    "R2": files.get("R2", ""),
                    "I1": files.get("I1", ""),
                    "I2": files.get("I2", ""),
                }
            )
    return runs


def collect_processed_files(processed_dir: Path, sample_title: str) -> list[tuple[str, Path]]:
    """Return (geo_file_name, path) for each processed DGE file of a sample.

    Every sample's files share the same basenames (count_matrix.mtx, etc.), so
    the GEO file name is prefixed with the sample title to make it unique across
    the submission, as GEO requires.
    """
    dge_dir = processed_dir / sample_title / "DGE_filtered"
    result: list[tuple[str, Path]] = []
    for fname in ("count_matrix.mtx", "all_genes.csv", "cell_metadata.csv"):
        path = dge_dir / fname
        if path.exists():
            result.append((f"{sample_title}_{fname}", path))
    return result


def md5sum(path: Path) -> str:
    """Return the hex MD5 digest of a file, read in chunks."""
    digest = hashlib.md5()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stage_processed_files(metadata: list[dict], dest_dir: Path) -> int:
    """Copy each sample's processed files into dest_dir under their GEO names.

    The GEO (sample-prefixed) file name matches what is written to the SAMPLES
    and MD5 Checksums sheets, so the staged folder can be uploaded as-is.
    Returns the number of files copied.
    """
    dest_dir.mkdir(parents=True, exist_ok=True)
    count = 0
    for row in metadata:
        for name, path in collect_processed_files(PROCESSED_DIR, row["sample_title"]):
            shutil.copy2(path, dest_dir / name)
            count += 1
    return count


def cell_type_to_tissue(cell_type: str) -> str:
    mapping = {
        "hESC": "embryoid body",
        "Osteoprogenitor": "bone marrow",
        "Osteoblast": "bone",
    }
    return mapping.get(cell_type, "hESC-derived osteogenic lineage")


def build_sample_title_display(row: dict) -> str:
    """Human-readable title for a GEO sample record."""
    parts = [row["sample_title"]]
    if row["treatment"] != "UNT":
        parts.append(f"{row['treatment']} {row['treatment_level']}")
    parts.append(f"day {row['day']}")
    return ", ".join(parts)


TREATMENT_PROTOCOL = (
    "Cells were treated with bisphenol analogues (BPA, BPF, BPS) at IC10, "
    "IC25, or IC50 concentrations, or with Zyn nicotine pouch extracts "
    "(Chill, Wintergreen, Smooth, Original) at 3 mg or 6 mg equivalents, "
    "for 24 hours prior to harvest. Untreated controls received vehicle only."
)

# Anchor rows within the template's "Metadata" sheet (1-indexed). The STUDY
# section and the SAMPLES header are fixed; SAMPLES data begins at
# SAMPLE_FIRST_ROW and the PROTOCOLS/PAIRED-END sections that follow are shifted
# down when more sample rows than the template provides are needed.
STUDY_ROWS = {
    "title": 12,
    "summary": 13,
    "experimental design": 14,
    "contributor": 15,
    "supplementary file": 22,
}
SAMPLE_HEADER_ROW = 38
SAMPLE_FIRST_ROW = 39
SAMPLE_TEMPLATE_ROWS = 15  # blank sample rows the template provides (39-53)
RAW_FIRST_COL = 18  # "*raw file" column after inserting a 3rd processed column (R1 pool; R2 pool in next)
PROTOCOL_ROWS = {
    "growth protocol": 57,
    "treatment protocol": 58,
    "extract protocol": 59,
    "library construction protocol": 60,
    "data processing step": 62,  # first of several consecutive rows
    "genome build/assembly": 67,
    "processed data files format and content": 68,
}
PAIRED_END_FIRST_ROW = 77
MD5_FIRST_ROW = 9  # first data row of the "MD5 Checksums" sheet


def build_geo_workbook(metadata: list[dict], fastq_runs: list[dict], title: str) -> openpyxl.Workbook:
    """Fill the official GEO template in place with this experiment's data."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Metadata"]

    _fill_metadata_sheet(ws, metadata, fastq_runs, title)
    _fill_md5_sheet(wb["MD5 Checksums"], fastq_runs, metadata)
    return wb


def _fill_metadata_sheet(ws, metadata: list[dict], fastq_runs: list[dict], title: str) -> int:
    # STUDY section
    ws.cell(row=STUDY_ROWS["title"], column=2, value=title)
    ws.cell(row=STUDY_ROWS["summary"], column=2, value=STUDY_SUMMARY_PLACEHOLDER)
    ws.cell(row=STUDY_ROWS["experimental design"], column=2, value=EXPERIMENTAL_DESIGN_PLACEHOLDER)
    ws.cell(row=STUDY_ROWS["contributor"], column=2, value="Sparks, Nicole")
    ws.cell(row=STUDY_ROWS["contributor"] + 1, column=2,
            value="[TODO: Add additional contributors as Lastname, Firstname]")

    # SAMPLES: the template ships two "processed data file" columns (O, P) but
    # each sample has three processed outputs, so add a third column. Insert it
    # to the left of the raw-file columns, which keeps the required-field
    # dropdowns (columns C, D, K, L, M) in place.
    ws.insert_cols(17)
    ws.cell(row=SAMPLE_HEADER_ROW, column=17, value="processed data file")

    # Every biological sample is spread across all Parse sublibraries via
    # combinatorial barcoding, so each sample's raw files are the full shared
    # FASTQ pool. GEO supports this: rather than one column per file, list each
    # read type's pooled files in its own "*raw file"/"raw file" cell,
    # comma-separated, identical on every sample row. The template provides four
    # raw-file columns, filled here with the pooled R1, R2, I1, and I2 lists. The
    # per-file R1/R2/I1/I2 grouping is enumerated once in the PAIRED-END
    # EXPERIMENTS section.
    r1_pool = ", ".join(run["R1"] for run in fastq_runs if run["R1"])
    r2_pool = ", ".join(run["R2"] for run in fastq_runs if run["R2"])
    i1_pool = ", ".join(run["I1"] for run in fastq_runs if run["I1"])
    i2_pool = ", ".join(run["I2"] for run in fastq_runs if run["I2"])

    # Make room for samples beyond the template's blank rows so the inserted
    # rows land above the PROTOCOLS section.
    offset = max(0, len(metadata) - SAMPLE_TEMPLATE_ROWS)
    if offset:
        ws.insert_rows(SAMPLE_FIRST_ROW + SAMPLE_TEMPLATE_ROWS, offset)

    for i, row in enumerate(metadata):
        r = SAMPLE_FIRST_ROW + i
        proc_names = [name for name, _ in collect_processed_files(PROCESSED_DIR, row["sample_title"])]
        while len(proc_names) < 3:
            proc_names.append("")
        treatment = "untreated" if row["treatment"] == "UNT" \
            else f"{row['treatment']} {row['treatment_level']}"
        values = [
            row["sample_title"],                    # A *library name
            build_sample_title_display(row),        # B *title (incl. level + day)
            LIBRARY_STRATEGY,                       # C *library strategy
            ORGANISM,                               # D *organism
            cell_type_to_tissue(row["cell_type"]),  # E **tissue
            "hESC-derived",                         # F **cell line
            row["cell_type"],                       # G **cell type
            "",                                     # H genotype
            treatment,                              # I treatment
            "",                                     # J batch
            MOLECULE,                               # K *molecule
            SINGLE_OR_PAIRED,                       # L *single or paired-end
            INSTRUMENT_MODEL,                       # M *instrument model
            row["experiment_explanation"],          # N description
            proc_names[0],                          # O processed data file
            proc_names[1],                          # P processed data file
            proc_names[2],                          # Q processed data file
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=value)
        # Raw files: the same pooled R1/R2/I1/I2 FASTQ lists on every sample row.
        ws.cell(row=r, column=RAW_FIRST_COL, value=r1_pool)      # R *raw file (R1 pool)
        ws.cell(row=r, column=RAW_FIRST_COL + 1, value=r2_pool)  # S raw file (R2 pool)
        ws.cell(row=r, column=RAW_FIRST_COL + 2, value=i1_pool)  # T raw file (I1 pool)
        ws.cell(row=r, column=RAW_FIRST_COL + 3, value=i2_pool)  # U raw file (I2 pool)

    # PROTOCOLS section (shifted down by the inserted sample rows)
    ws.cell(row=PROTOCOL_ROWS["treatment protocol"] + offset, column=2, value=TREATMENT_PROTOCOL)
    ws.cell(row=PROTOCOL_ROWS["extract protocol"] + offset, column=2, value=EXTRACT_PROTOCOL)
    ws.cell(row=PROTOCOL_ROWS["library construction protocol"] + offset, column=2,
            value=LIBRARY_CONSTRUCTION_PROTOCOL)
    for i, step in enumerate(DATA_PROCESSING_STEPS):
        ws.cell(row=PROTOCOL_ROWS["data processing step"] + offset + i, column=2, value=step)
    ws.cell(row=PROTOCOL_ROWS["genome build/assembly"] + offset, column=2, value=GENOME_BUILD)
    ws.cell(row=PROTOCOL_ROWS["processed data files format and content"] + offset, column=2,
            value=PROCESSED_FILES_FORMAT)

    # PAIRED-END EXPERIMENTS section (R1/R2/I1/I2 per FASTQ run)
    for i, run in enumerate(fastq_runs):
        r = PAIRED_END_FIRST_ROW + offset + i
        for c, key in enumerate(("R1", "R2", "I1", "I2"), start=1):
            ws.cell(row=r, column=c, value=run[key])

    return offset


def _fill_md5_sheet(ws, fastq_runs: list[dict], metadata: list[dict]):
    checksums = load_checksums(CHECKSUMS_FILE)

    # RAW FILES (columns A/B): checksums come from the precomputed md5 file.
    r = MD5_FIRST_ROW
    for run in fastq_runs:
        for key in ("R1", "R2", "I1", "I2"):
            fname = run[key]
            if fname:
                ws.cell(row=r, column=1, value=fname)
                ws.cell(row=r, column=2, value=checksums.get(fname, "[TODO: compute]"))
                r += 1

    # PROCESSED DATA FILES (columns F/G): computed on the fly from the files on
    # disk, using the same sample-prefixed names written to the SAMPLES section.
    r = MD5_FIRST_ROW
    for row in metadata:
        for name, path in collect_processed_files(PROCESSED_DIR, row["sample_title"]):
            ws.cell(row=r, column=6, value=name)
            ws.cell(row=r, column=7, value=md5sum(path))
            r += 1


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR,
        help="Output directory for the per-experiment GEO workbooks (default: output/)"
    )
    parser.add_argument(
        "--no-stage", action="store_true",
        help="Skip copying processed files into per-experiment upload folders"
    )
    args = parser.parse_args()

    if not METADATA_CSV.exists():
        sys.exit(f"ERROR: metadata not found at {METADATA_CSV}")
    if not FASTQ_DIR.exists():
        sys.exit(f"ERROR: FASTQ directory not found at {FASTQ_DIR}")
    if not TEMPLATE_PATH.exists():
        sys.exit(f"ERROR: GEO template not found at {TEMPLATE_PATH}")

    print(f"Loading submission titles from {CONFIG_PATH}")
    titles = load_submission_titles(CONFIG_PATH)
    if not titles:
        sys.exit(f"ERROR: no geo.submissions titles found in {CONFIG_PATH}")

    print(f"Loading metadata from {METADATA_CSV}")
    metadata = load_metadata(METADATA_CSV)
    print(f"  {len(metadata)} samples")

    print(f"Scanning FASTQ directory {FASTQ_DIR}")
    fastq_runs = collect_fastq_runs(FASTQ_DIR)
    print(f"  {len(fastq_runs)} FASTQ runs (sublibrary × lane)")

    sample_list = load_sample_list(SAMPLE_LIST_FILE) if SAMPLE_LIST_FILE.exists() else []
    if not sample_list:
        print(f"  NOTE: {SAMPLE_LIST_FILE} not found; skipping per-experiment sample lists")

    args.output_dir.mkdir(parents=True, exist_ok=True)

    experiments = sorted({row["experiment"] for row in metadata})
    for exp in experiments:
        exp_metadata = [row for row in metadata if row["experiment"] == exp]
        title = titles.get(exp)
        if title is None:
            sys.exit(f"ERROR: no title configured for experiment {exp} in {CONFIG_PATH}")

        print(f"\nBuilding GEO workbook for experiment {exp} ({len(exp_metadata)} samples)...")
        print(f"  title: {title}")
        wb = build_geo_workbook(exp_metadata, fastq_runs, title)

        output = args.output_dir / f"geo_submission_metadata_exp{exp}.xlsx"
        wb.save(output)
        print(f"  saved to {output}")

        if not args.no_stage:
            upload_dir = args.output_dir / f"geo_upload_exp{exp}"
            n_staged = stage_processed_files(exp_metadata, upload_dir)
            print(f"  staged {n_staged} processed files to {upload_dir}")

        if sample_list:
            exp_titles = {row["sample_title"] for row in exp_metadata}
            exp_lines = [line for name, line in sample_list if name in exp_titles]
            missing = exp_titles - {name for name, _ in sample_list}
            if missing:
                print(f"  WARNING: no sample_list entry for: {', '.join(sorted(missing))}")
            sample_list_out = args.output_dir / f"sample_list_exp{exp}.txt"
            sample_list_out.write_text("\n".join(exp_lines) + "\n")
            print(f"  wrote {len(exp_lines)} sample_list entries to {sample_list_out}")

    # Summary of what needs manual review
    print("\n--- Fields requiring manual review ---")
    print("  STUDY sheet:")
    print("    *summary (abstract)  — insert final abstract text")
    print("    contributor          — add all co-authors")
    print("    supplementary file   — add path to combined AnnData/Seurat object if any")
    print("  SAMPLES section:")
    print("    raw file 1-4         — pooled R1, R2, I1, I2 FASTQ lists (comma-separated),")
    print("                           identical on every sample row (shared Parse pool)")
    print("  PROTOCOLS section:")
    print("    growth protocol      — add hESC culture and differentiation protocol details")
    print("  Upload:")
    print("    geo_upload_exp*/     — processed files staged under GEO names; upload these")
    print("    raw FASTQs           — upload from data/FASTQ (names already match the sheet)")


if __name__ == "__main__":
    main()
